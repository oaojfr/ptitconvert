"""
Convertisseur de feuilles de calcul pour PtitConvert
Gère la conversion entre XLSX, CSV et PDF
"""

import openpyxl
import csv
import pandas as pd
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import os
from pathlib import Path

class SpreadsheetConverter:
    """Convertisseur pour les feuilles de calcul"""
    
    SUPPORTED_INPUT_FORMATS = {'.xlsx', '.csv'}
    SUPPORTED_OUTPUT_FORMATS = {'xlsx', 'csv', 'pdf'}
    
    def __init__(self):
        """Initialiser le convertisseur de feuilles de calcul"""
        pass
        
    def convert(self, input_path, output_dir, output_format):
        """
        Convertir une feuille de calcul vers le format spécifié
        
        Args:
            input_path (str): Chemin du fichier d'entrée
            output_dir (str): Répertoire de sortie
            output_format (str): Format de sortie ('xlsx', 'csv', 'pdf')
            
        Returns:
            bool: True si la conversion a réussi, False sinon
        """
        try:
            input_path = Path(input_path)
            output_format = output_format.lower()
            
            # Vérifier que le format d'entrée est supporté
            if input_path.suffix.lower() not in self.SUPPORTED_INPUT_FORMATS:
                print(f"Format d'entrée non supporté: {input_path.suffix}")
                return False
                
            # Vérifier que le format de sortie est supporté
            if output_format not in self.SUPPORTED_OUTPUT_FORMATS:
                print(f"Format de sortie non supporté: {output_format}")
                return False
                
            # Créer le nom de fichier de sortie
            output_name = f"{input_path.stem}.{output_format}"
            output_path = Path(output_dir) / output_name
            
            # Lire les données du fichier source
            data = self._read_data(input_path)
            if data is None:
                return False
                
            # Convertir selon le format de sortie
            if output_format == 'xlsx':
                return self._create_xlsx(data, output_path)
            elif output_format == 'csv':
                return self._create_csv(data, output_path)
            elif output_format == 'pdf':
                return self._create_pdf(data, output_path)
            else:
                return False
                
        except Exception as e:
            print(f"Erreur lors de la conversion de feuille de calcul: {e}")
            return False
            
    def _read_data(self, input_path):
        """
        Lire les données d'une feuille de calcul
        
        Args:
            input_path (Path): Chemin du fichier d'entrée
            
        Returns:
            list: Données sous forme de liste de listes ou None en cas d'erreur
        """
        try:
            file_ext = input_path.suffix.lower()
            
            if file_ext == '.xlsx':
                return self._read_xlsx(input_path)
            elif file_ext == '.csv':
                return self._read_csv(input_path)
            else:
                return None
                
        except Exception as e:
            print(f"Erreur lors de la lecture des données: {e}")
            return None
            
    def _read_xlsx(self, xlsx_path):
        """Lire un fichier Excel"""
        try:
            workbook = openpyxl.load_workbook(xlsx_path)
            worksheet = workbook.active
            
            data = []
            for row in worksheet.iter_rows(values_only=True):
                # Convertir None en chaîne vide et tout en string
                row_data = [str(cell) if cell is not None else '' for cell in row]
                data.append(row_data)
                
            return data
            
        except Exception as e:
            print(f"Erreur lors de la lecture du XLSX: {e}")
            return None
            
    def _read_csv(self, csv_path):
        """Lire un fichier CSV"""
        try:
            data = []
            
            # Essayer différents encodages
            encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
            
            for encoding in encodings:
                try:
                    with open(csv_path, 'r', encoding=encoding, newline='') as file:
                        # Détecter le délimiteur
                        sample = file.read(1024)
                        file.seek(0)
                        
                        sniffer = csv.Sniffer()
                        delimiter = sniffer.sniff(sample).delimiter
                        
                        reader = csv.reader(file, delimiter=delimiter)
                        for row in reader:
                            data.append(row)
                            
                    return data
                    
                except UnicodeDecodeError:
                    continue
                    
            print(f"Impossible de lire le fichier CSV avec les encodages supportés")
            return None
            
        except Exception as e:
            print(f"Erreur lors de la lecture du CSV: {e}")
            return None
            
    def _create_xlsx(self, data, output_path):
        """Créer un fichier Excel à partir des données"""
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
                    
            workbook.save(str(output_path))
            
            print(f"XLSX créé: {output_path}")
            return True
            
        except Exception as e:
            print(f"Erreur lors de la création du XLSX: {e}")
            return False
            
    def _create_csv(self, data, output_path):
        """Créer un fichier CSV à partir des données"""
        try:
            with open(output_path, 'w', encoding='utf-8', newline='') as file:
                writer = csv.writer(file)
                for row in data:
                    writer.writerow(row)
                    
            print(f"CSV créé: {output_path}")
            return True
            
        except Exception as e:
            print(f"Erreur lors de la création du CSV: {e}")
            return False
            
    def _create_pdf(self, data, output_path):
        """Créer un PDF à partir des données"""
        try:
            doc = SimpleDocTemplate(str(output_path), pagesize=A4)
            
            # Limiter le nombre de colonnes et de lignes pour l'affichage
            max_cols = 8
            max_rows = 50
            
            # Tronquer les données si nécessaire
            display_data = []
            for i, row in enumerate(data):
                if i >= max_rows:
                    break
                    
                # Tronquer la ligne si trop de colonnes
                if len(row) > max_cols:
                    display_row = row[:max_cols-1] + ['...']
                else:
                    display_row = row
                    
                # Limiter la longueur du contenu des cellules
                display_row = [str(cell)[:20] + '...' if len(str(cell)) > 20 else str(cell) 
                              for cell in display_row]
                display_data.append(display_row)
                
            if len(data) > max_rows:
                display_data.append(['...'] * len(display_data[0]))
                
            # Créer le tableau
            table = Table(display_data)
            
            # Style du tableau
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            # Construire le document
            elements = [table]
            doc.build(elements)
            
            print(f"PDF créé: {output_path}")
            return True
            
        except Exception as e:
            print(f"Erreur lors de la création du PDF: {e}")
            return False
            
    def get_spreadsheet_info(self, spreadsheet_path):
        """
        Obtenir les informations d'une feuille de calcul
        
        Args:
            spreadsheet_path (str): Chemin de la feuille de calcul
            
        Returns:
            dict: Informations sur la feuille de calcul
        """
        try:
            path = Path(spreadsheet_path)
            file_ext = path.suffix.lower()
            
            info = {
                'file_size': path.stat().st_size,
                'file_format': file_ext[1:].upper()
            }
            
            if file_ext == '.xlsx':
                workbook = openpyxl.load_workbook(spreadsheet_path)
                worksheet = workbook.active
                info['row_count'] = worksheet.max_row
                info['column_count'] = worksheet.max_column
                info['sheet_names'] = workbook.sheetnames
                
            elif file_ext == '.csv':
                with open(spreadsheet_path, 'r', encoding='utf-8') as file:
                    reader = csv.reader(file)
                    rows = list(reader)
                    info['row_count'] = len(rows)
                    info['column_count'] = len(rows[0]) if rows else 0
                    
            return info
            
        except Exception as e:
            print(f"Erreur lors de la lecture des informations de la feuille de calcul: {e}")
            return None
            
    def convert_to_dataframe(self, file_path):
        """
        Convertir un fichier en DataFrame pandas pour manipulation avancée
        
        Args:
            file_path (str): Chemin du fichier
            
        Returns:
            pandas.DataFrame: DataFrame ou None en cas d'erreur
        """
        try:
            path = Path(file_path)
            file_ext = path.suffix.lower()
            
            if file_ext == '.xlsx':
                return pd.read_excel(file_path)
            elif file_ext == '.csv':
                return pd.read_csv(file_path)
            else:
                return None
                
        except Exception as e:
            print(f"Erreur lors de la conversion en DataFrame: {e}")
            return None
